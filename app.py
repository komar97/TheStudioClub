import base64
import os
import sqlite3
from io import BytesIO
from pathlib import Path
from datetime import date, datetime

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from sqlalchemy import create_engine, text
from sqlalchemy.exc import IntegrityError as SQLAlchemyIntegrityError


st.set_page_config(
    page_title="The Studio Club - Stock",
    page_icon="🍾",
    layout="wide",
    initial_sidebar_state="collapsed",
)


# Par defaut, l'app utilise SQLite en local.
# En production, definir DATABASE_URL dans .streamlit/secrets.toml
# ou comme variable d'environnement pour utiliser une base PostgreSQL persistante
# Ex : DATABASE_URL = "postgresql+psycopg2://user:password@host:5432/postgres"
DB_FILE = Path("stock.db")


def get_database_url():
    try:
        return st.secrets.get("DATABASE_URL", None) or st.secrets.get("database", {}).get("url", None)
    except Exception:
        return None


@st.cache_resource
def get_engine():
    database_url = get_database_url() or os.getenv("DATABASE_URL")
    if database_url:
        return create_engine(
            database_url,
            pool_pre_ping=True,
            pool_size=3,
            max_overflow=2,
            pool_recycle=1800,
        )
    return create_engine(f"sqlite:///{DB_FILE}", connect_args={"check_same_thread": False})


ENGINE = get_engine()
IS_POSTGRES = ENGINE.url.get_backend_name().startswith("postgresql")
ADMIN_PASSWORD = "1234"
STOCK_PASSWORD = "01234"

ARTICLES_NON_ALCOOL = [
    "JUS", "PEPSI", "LIMONADE", "TONIC", "MEGA FORCE", "PERRIER",
    "ICE TEA", "GINGER BEER", "BIB PEPSI", "BIB TONIC", "BIB HEXIS",
    "BIB 7UP", "SIROP", "PULCO", "LAIT"
]


INITIAL_ARTICLES = [
    ('AMARETTO', 15),
    ('BAMBU', 4),
    ('BELVEDERE 1,75L', 3),
    ('BELVEDERE 3L', 0),
    ('BELVEDERE 70 CL', 5),
    ('BIB 7UP', 2),
    ('BIB HEXIS', 0),
    ('BIB PEPSI', 2),
    ('BIB TONIC', 4),
    ('CAPTAIN MORGAN', 49),
    ('CHAMPAGNE DEUTZ 1,5L', 2),
    ('CHAMPAGNE DEUTZ BLANC', 4),
    ('CHAMPAGNE DON PERIGNON', 1),
    ('CHAMPAGNE LA GRANDE DAME', 2),
    ('CHAMPAGNE MOET & CHANDON', 12),
    ('CHAMPAGNE RUINART BRUT', 0),
    ('CIROC PEACH', 14),
    ('CIROC RED BERRY', 8),
    ('CLAN CAMPBELL', 24),
    ('COINTREAU', 16),
    ('FJOROKWA 3L', -1),
    ('FJOROKWA 70CL', 33),
    ('GET 27', 25),
    ('GIN GRAND CABARET', 0),
    ('GIN HENDRICKS', 0),
    ('GIN LION HEART', 30),
    ('IRISH CREAM', 18),
    ('JACK FIRE', 9),
    ('JACK N7', 7),
    ('JAGGERMEISTER', 47),
    ("JUS D'ANANAS", 27),
    ("JUS D'ORANGE", 32),
    ('JUS DE POMME', 24),
    ('LIMONADE', 85),
    ('LIQUEUR DE CAFE', 8),
    ('MEGA FORCE / CRAZY TIGER', 76),
    ('MONKEY SHOULDERS', 1),
    ('PEPPERMINT 5L', 0),
    ('PEPSI', 73),
    ('PERRIER', 32),
    ('PICON', 8),
    ('RICARD', 23),
    ('SIROP CITRON', 3),
    ('SIROP GRENADINE', 6),
    ('SIROP MENTHE', 3),
    ('SIROP ORGEAT', 7),
    ('SIROP PECHE', 13),
    ('SMIRNOFF', 131),
    ('TEQUILA ACAYUCAN', 2),
    ('TONIC', 32),
]


def ensure_storage_dirs():
    # Plus de sauvegarde OneDrive : les rapports sont telecharges par l'utilisateur.
    return None


def connect():
    # Compatibilite avec les anciennes fonctions si besoin.
    return sqlite3.connect(DB_FILE)


def init_db():
    if IS_POSTGRES:
        articles_sql = """
            CREATE TABLE IF NOT EXISTS articles (
                id SERIAL PRIMARY KEY,
                nom TEXT NOT NULL UNIQUE,
                stock INTEGER NOT NULL DEFAULT 0
            )
        """
        mouvements_sql = """
            CREATE TABLE IF NOT EXISTS mouvements (
                id SERIAL PRIMARY KEY,
                date_jour TEXT NOT NULL,
                heure TEXT NOT NULL,
                personne TEXT NOT NULL,
                article_id INTEGER NOT NULL,
                article_nom TEXT NOT NULL,
                sortie INTEGER NOT NULL DEFAULT 0,
                entree INTEGER NOT NULL DEFAULT 0,
                ancien_stock INTEGER NOT NULL,
                nouveau_stock INTEGER NOT NULL
            )
        """
    else:
        articles_sql = """
            CREATE TABLE IF NOT EXISTS articles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nom TEXT NOT NULL UNIQUE,
                stock INTEGER NOT NULL DEFAULT 0
            )
        """
        mouvements_sql = """
            CREATE TABLE IF NOT EXISTS mouvements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date_jour TEXT NOT NULL,
                heure TEXT NOT NULL,
                personne TEXT NOT NULL,
                article_id INTEGER NOT NULL,
                article_nom TEXT NOT NULL,
                sortie INTEGER NOT NULL DEFAULT 0,
                entree INTEGER NOT NULL DEFAULT 0,
                ancien_stock INTEGER NOT NULL,
                nouveau_stock INTEGER NOT NULL
            )
        """

    with ENGINE.begin() as conn:
        conn.execute(text(articles_sql))
        conn.execute(text(mouvements_sql))
        count = conn.execute(text("SELECT COUNT(*) FROM articles")).scalar()
        if count == 0:
            conn.execute(
                text("INSERT INTO articles (nom, stock) VALUES (:nom, :stock)"),
                [{"nom": nom, "stock": stock} for nom, stock in INITIAL_ARTICLES],
            )


@st.cache_resource
def initialize_database():
    init_db()
    return True


@st.cache_data(ttl=30, show_spinner=False)
def get_articles():
    with ENGINE.connect() as conn:
        return pd.read_sql_query(
            text("""
                SELECT id, nom AS article, stock
                FROM articles
                ORDER BY LOWER(nom)
            """),
            conn,
        )


def clear_data_cache():
    get_articles.clear()
    get_mouvements.clear()
    get_daily_summary.clear()


def update_stock(article_id, stock):
    with ENGINE.begin() as conn:
        conn.execute(
            text("UPDATE articles SET stock = :stock WHERE id = :id"),
            {"stock": int(stock), "id": int(article_id)},
        )
    clear_data_cache()


def add_article(nom, stock):
    with ENGINE.begin() as conn:
        conn.execute(
            text("INSERT INTO articles (nom, stock) VALUES (:nom, :stock)"),
            {"nom": nom.strip(), "stock": int(stock)},
        )
    clear_data_cache()


def delete_article(article_id):
    with ENGINE.begin() as conn:
        conn.execute(text("DELETE FROM articles WHERE id = :id"), {"id": int(article_id)})
    clear_data_cache()


def add_movement(personne, article_id, article_nom, sortie, entree, ancien_stock, nouveau_stock):
    now = datetime.now()
    with ENGINE.begin() as conn:
        conn.execute(
            text("""
                INSERT INTO mouvements (
                    date_jour, heure, personne, article_id, article_nom,
                    sortie, entree, ancien_stock, nouveau_stock
                )
                VALUES (
                    :date_jour, :heure, :personne, :article_id, :article_nom,
                    :sortie, :entree, :ancien_stock, :nouveau_stock
                )
            """),
            {
                "date_jour": now.strftime("%Y-%m-%d"),
                "heure": now.strftime("%H:%M:%S"),
                "personne": personne,
                "article_id": int(article_id),
                "article_nom": article_nom,
                "sortie": int(sortie),
                "entree": int(entree),
                "ancien_stock": int(ancien_stock),
                "nouveau_stock": int(nouveau_stock),
            },
        )
    clear_data_cache()


def apply_movements(personne, mouvements_valides):
    if not mouvements_valides:
        return

    now = datetime.now()
    movement_rows = []
    stock_rows = []

    for move in mouvements_valides:
        movement_rows.append({
            "date_jour": now.strftime("%Y-%m-%d"),
            "heure": now.strftime("%H:%M:%S"),
            "personne": personne,
            "article_id": int(move["id"]),
            "article_nom": move["article"],
            "sortie": int(move["sortie"]),
            "entree": int(move["entree"]),
            "ancien_stock": int(move["ancien_stock"]),
            "nouveau_stock": int(move["nouveau_stock"]),
        })
        stock_rows.append({
            "id": int(move["id"]),
            "stock": int(move["nouveau_stock"]),
        })

    with ENGINE.begin() as conn:
        conn.execute(
            text("UPDATE articles SET stock = :stock WHERE id = :id"),
            stock_rows,
        )
        conn.execute(
            text("""
                INSERT INTO mouvements (
                    date_jour, heure, personne, article_id, article_nom,
                    sortie, entree, ancien_stock, nouveau_stock
                )
                VALUES (
                    :date_jour, :heure, :personne, :article_id, :article_nom,
                    :sortie, :entree, :ancien_stock, :nouveau_stock
                )
            """),
            movement_rows,
        )

    clear_data_cache()


@st.cache_data(ttl=60, show_spinner=False)
def get_mouvements():
    with ENGINE.connect() as conn:
        return pd.read_sql_query(
            text("""
                SELECT date_jour, heure, personne, article_nom AS article,
                       sortie, entree, ancien_stock, nouveau_stock
                FROM mouvements
                ORDER BY date_jour DESC, heure DESC, id DESC
            """),
            conn,
        )

@st.cache_data(ttl=60, show_spinner=False)
def get_daily_summary():
    with ENGINE.connect() as conn:
        return pd.read_sql_query(
            text("""
                SELECT date_jour,
                       SUM(sortie) AS sorties,
                       SUM(entree) AS entrees
                FROM mouvements
                GROUP BY date_jour
                ORDER BY date_jour ASC
            """),
            conn,
        )


def is_alcool(article):
    nom = str(article).strip()
    return not any(nom.lower().startswith(prefix.lower()) for prefix in ARTICLES_NON_ALCOOL)


def style_excel_sheet(ws):
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        width = min(max(max_len + 3, 12), 38)
        ws.column_dimensions[get_column_letter(col[0].column)].width = width


def is_stock_faible(stock):
    try:
        return int(stock) <= int(st.session_state.get("seuil_stock_faible", 3))
    except Exception:
        return False


def build_excel_report(personne, mouvements_valides):
    from openpyxl.styles import Font, PatternFill

    stock_actuel = get_articles()

    # L'historique contient TOUS les articles du stock, pas seulement les alcools.
    stock_export = stock_actuel.copy().rename(columns={"article": "Article", "stock": "Stock actuel"})
    stock_export["Catégorie"] = stock_export["Article"].apply(lambda x: "Alcool" if is_alcool(x) else "Soft / autre")
    stock_export["Alerte"] = stock_export["Stock actuel"].apply(lambda x: "Stock faible" if is_stock_faible(x) else "")
    stock_export = stock_export[["Article", "Catégorie", "Stock actuel", "Alerte"]]

    resume = pd.DataFrame(mouvements_valides)
    if not resume.empty:
        resume = resume.rename(columns={
            "article": "Article",
            "sortie": "Sortie",
            "entree": "Entrée",
            "ancien_stock": "Ancien stock",
            "nouveau_stock": "Nouveau stock",
        })
        resume = resume[["Article", "Sortie", "Entrée", "Ancien stock", "Nouveau stock"]]
    else:
        resume = pd.DataFrame(columns=["Article", "Sortie", "Entrée", "Ancien stock", "Nouveau stock"])

    totaux = pd.DataFrame([{
        "Date": date.today().strftime("%d/%m/%Y"),
        "Heure": datetime.now().strftime("%H:%M:%S"),
        "Personne": personne,
        "Total sorties": int(resume["Sortie"].sum()) if not resume.empty else 0,
        "Total entrées": int(resume["Entrée"].sum()) if not resume.empty else 0,
        "Articles modifiés": len(resume),
        "Articles au total": len(stock_export),
        "Stock total": int(stock_export["Stock actuel"].sum()) if not stock_export.empty else 0,
        "Articles en stock faible": int((stock_export["Alerte"] == "Stock faible").sum()) if not stock_export.empty else 0,
    }])

    filename = datetime.now().strftime("The_Studio_Club_Stock_%Y-%m-%d_%H-%M-%S.xlsx")
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        stock_export.to_excel(writer, sheet_name="Stock actuel", index=False)
        resume.to_excel(writer, sheet_name="Mouvements du jour", index=False)
        totaux.to_excel(writer, sheet_name="Résumé", index=False)

        for sheet_name in writer.sheets:
            style_excel_sheet(writer.sheets[sheet_name])

        red_fill = PatternFill("solid", fgColor="FEE2E2")
        red_font = Font(color="991B1B", bold=True)
        ws_stock = writer.sheets["Stock actuel"]
        alerte_col = ws_stock.max_column
        for row in range(2, ws_stock.max_row + 1):
            if ws_stock.cell(row=row, column=alerte_col).value == "Stock faible":
                for col in range(1, ws_stock.max_column + 1):
                    ws_stock.cell(row=row, column=col).fill = red_fill
                    ws_stock.cell(row=row, column=col).font = red_font

    return filename, output.getvalue()


def prepare_download_report(personne, mouvements_valides):
    filename, excel_bytes = build_excel_report(personne, mouvements_valides)
    st.session_state.last_report_filename = filename
    st.session_state.last_report_bytes = excel_bytes
    st.session_state.auto_download_report = True


def render_report_download_box():
    filename = st.session_state.get("last_report_filename")
    excel_bytes = st.session_state.get("last_report_bytes")

    if not filename or not excel_bytes:
        return

    st.success("Rapport Excel généré. Télécharge-le pour garder l'historique du stock.")

    st.download_button(
        "📥 Télécharger le rapport Excel",
        data=excel_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    # Tentative de téléchargement automatique.
    # Selon l'iPad, Safari ou Chrome, le navigateur peut bloquer l'action.
    if st.session_state.get("auto_download_report"):
        b64 = base64.b64encode(excel_bytes).decode("utf-8")
        html = f'''
            <html>
                <body>
                    <a id="download_excel"
                       href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}"
                       download="{filename}">Télécharger</a>
                    <script>
                        const link = document.getElementById('download_excel');
                        if (link) {{ link.click(); }}
                    </script>
                </body>
            </html>
        '''
        components.html(html, height=0)
        st.session_state.auto_download_report = False


def movement_key(kind, article_id):
    return f"{kind}_{article_id}"


def ensure_movement_keys(article_id):
    for kind in ["sortie", "entree"]:
        key = movement_key(kind, article_id)
        if key not in st.session_state:
            st.session_state[key] = 0


def change_movement(kind, article_id, delta):
    key = movement_key(kind, article_id)
    current = int(st.session_state.get(key, 0) or 0)
    st.session_state[key] = max(0, current + delta)


def reset_movements():
    for key in list(st.session_state.keys()):
        if key.startswith("sortie_") or key.startswith("entree_"):
            del st.session_state[key]


def css():
    st.markdown("""
        <style>
            .stApp {
                background: #070b14;
                color: #ffffff;
            }

            .block-container {
                padding-top: 3.2rem;
                max-width: 1280px;
            }

            h1 {
                font-size: 34px !important;
                font-weight: 950 !important;
                letter-spacing: -0.5px;
                margin-bottom: 0.2rem !important;
                line-height: 1.25 !important;
            }

            .app-title {
                font-size: 38px;
                line-height: 1.35;
                font-weight: 950;
                color: #ffffff;
                letter-spacing: -0.5px;
                margin-bottom: 6px;
                padding-top: 4px;
            }

            h2, h3 {
                color: #ffffff;
                font-weight: 850 !important;
            }

            label {
                font-weight: 750 !important;
                color: #e5e7eb !important;
            }

            input {
                min-height: 40px !important;
                font-size: 16px !important;
                border-radius: 10px !important;
            }

            button {
                border-radius: 12px !important;
                font-weight: 900 !important;
                min-height: 44px !important;
                line-height: 1.2 !important;
                overflow: visible !important;
            }

            div[data-testid="stButton"] button {
                color: #ffffff !important;
            }

            div[data-testid="stMetric"] {
                background: #111827;
                border: 1px solid #263145;
                border-radius: 16px;
                padding: 14px 16px;
            }

            div[data-testid="stMetricValue"] {
                font-size: 28px !important;
            }

            .top-subtitle {
                color: #aab2c5;
                font-size: 15px;
                margin-bottom: 18px;
            }

            .date-pill {
                display: inline-block;
                background: #111827;
                border: 1px solid #334155;
                color: #f3f4f6;
                border-radius: 999px;
                padding: 9px 16px;
                font-size: 15px;
                font-weight: 850;
                margin-top: 16px;
                margin-bottom: 18px;
            }

            .nav-active button {
                border: 1.5px solid #ff4655 !important;
                background: #1a2232 !important;
                box-shadow: 0 0 0 1px rgba(255, 70, 85, 0.25) inset !important;
            }

            .nav-normal button {
                border: 1px solid #3a455c !important;
                background: #111827 !important;
            }

            .admin-normal button {
                background: #ff4655 !important;
                border: 1px solid #ff5a66 !important;
                margin-left: 18px !important;
            }

            .admin-active button {
                background: #ff4655 !important;
                border: 2px solid #ffd1d5 !important;
                box-shadow: 0 0 0 2px rgba(255, 70, 85, 0.30) inset !important;
                margin-left: 18px !important;
            }

            .table-head {
                background: #1d2638;
                border: 1px solid #33405a;
                border-radius: 14px;
                padding: 11px 14px;
                font-size: 18px;
                font-weight: 950;
                color: #ffffff;
                text-align: center;
            }

            .cell-a, .cell-b {
                border-radius: 14px;
                min-height: 54px;
                padding: 9px 12px;
                border: 1px solid #2a354c;
                display: flex;
                align-items: center;
            }

            .cell-a {
                background: #101827;
            }

            .cell-b {
                background: #172136;
            }

            .article {
                font-size: 19px;
                font-weight: 850;
                line-height: 1.15;
                color: #ffffff;
            }

            .stock {
                width: 100%;
                text-align: center;
                font-size: 20px;
                font-weight: 950;
                color: #f3f4f6;
            }

            .movement-value {
                background: #0b1220;
                border: 1px solid #334155;
                border-radius: 12px;
                min-height: 40px;
                display: flex;
                justify-content: center;
                align-items: center;
                font-size: 20px;
                font-weight: 950;
                color: #ffffff;
            }

            .row-gap {
                height: 5px;
            }

            .delete-note {
                background: #2a1216;
                color: #fecaca;
                border: 1px solid #7f1d1d;
                padding: 12px 14px;
                border-radius: 14px;
                font-weight: 750;
            }

            [data-testid="column"] {
                padding-top: 0 !important;
                padding-bottom: 0 !important;
            }

            [data-testid="stVerticalBlock"] {
                gap: 0.35rem !important;
            }

            [data-testid="stDataFrame"] div {
                font-size: 17px !important;
            }


            div[data-testid="stNumberInput"] input {
                text-align: center !important;
                font-weight: 950 !important;
                color: #ffffff !important;
            }

            .big-alert {
                background: #7f1d1d;
                color: #ffffff;
                border: 3px solid #fecaca;
                border-radius: 18px;
                padding: 24px 28px;
                margin: 18px 0 22px 0;
                font-size: 28px;
                line-height: 1.25;
                font-weight: 950;
                text-align: center;
                box-shadow: 0 0 0 4px rgba(248, 113, 113, 0.25);
            }

            .big-alert-sub {
                display: block;
                margin-top: 8px;
                font-size: 18px;
                font-weight: 850;
                color: #fee2e2;
            }
        </style>
    """, unsafe_allow_html=True)


def render_big_missing_name_alert():
    st.markdown(
        """
        <div class="big-alert">
            ⚠️ NOM OBLIGATOIRE
            <span class="big-alert-sub">Écris le nom de la personne qui complète avant de valider le stock.</span>
        </div>
        """,
        unsafe_allow_html=True,
    )


def movement_control(kind, article_id):
    ensure_movement_keys(article_id)
    key = movement_key(kind, article_id)

    minus, value, plus = st.columns([0.55, 1.15, 0.55], gap="small")

    with minus:
        st.button(
            "−",
            key=f"{key}_minus",
            on_click=change_movement,
            args=(kind, article_id, -1),
            use_container_width=True,
        )

    with value:
        if key not in st.session_state:
            st.session_state[key] = 0
        st.number_input(
            "Quantité",
            min_value=0,
            step=1,
            key=key,
            label_visibility="collapsed",
        )

    with plus:
        st.button(
            "＋",
            key=f"{key}_plus",
            on_click=change_movement,
            args=(kind, article_id, 1),
            use_container_width=True,
        )

    return int(st.session_state.get(key, 0) or 0)


def set_page(page):
    st.session_state.page = page


css()
ensure_storage_dirs()
initialize_database()

if "seuil_stock_faible" not in st.session_state:
    st.session_state.seuil_stock_faible = 3

if "page" not in st.session_state:
    st.session_state.page = "stock"

if "last_report_filename" not in st.session_state:
    st.session_state.last_report_filename = None
if "last_report_bytes" not in st.session_state:
    st.session_state.last_report_bytes = None
if "auto_download_report" not in st.session_state:
    st.session_state.auto_download_report = False
if "stock_unlocked" not in st.session_state:
    st.session_state.stock_unlocked = False
if "missing_name_alert" not in st.session_state:
    st.session_state.missing_name_alert = False

st.markdown('<div class="app-title">🍾 The Studio Club</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="top-subtitle">Gestion de stock simple pour les entrées et sorties du jour</div>',
    unsafe_allow_html=True
)

nav_stock_active = "nav-active" if st.session_state.page == "stock" else "nav-normal"
nav_resume_active = "nav-active" if st.session_state.page == "resume" else "nav-normal"
nav_admin_active = "admin-active" if st.session_state.page == "admin" else "admin-normal"

nav_left, nav_gap, nav_admin = st.columns([2.2, 0.35, 1.2], gap="medium")

with nav_left:
    n1, n2 = st.columns(2, gap="small")
    with n1:
        st.markdown(f'<div class="{nav_stock_active}">', unsafe_allow_html=True)
        st.button("Stock", on_click=set_page, args=("stock",), use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)
    with n2:
        st.markdown(f'<div class="{nav_resume_active}">', unsafe_allow_html=True)
        st.button("Stats", on_click=set_page, args=("resume",), use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

with nav_admin:
    st.markdown(f'<div class="{nav_admin_active}">', unsafe_allow_html=True)
    st.button("🔒 Admin", on_click=set_page, args=("admin",), use_container_width=True)
    st.markdown("</div>", unsafe_allow_html=True)

today_label = date.today().strftime("%d/%m/%Y")
st.markdown(f'<div class="date-pill">📅 Date du jour : {today_label}</div>', unsafe_allow_html=True)

st.divider()


if st.session_state.page == "stock":
    if not st.session_state.stock_unlocked:
        st.subheader("Accès stock")
        stock_password = st.text_input("Mot de passe stock", type="password", placeholder="Mot de passe")
        if stock_password == STOCK_PASSWORD:
            st.session_state.stock_unlocked = True
            st.rerun()
        elif stock_password:
            st.error("Mot de passe incorrect.")
        st.stop()

    articles = get_articles()
    seuil = st.session_state.seuil_stock_faible

    col1, col2 = st.columns(2)
    col1.metric("Articles", len(articles))
    col2.metric("Stock total", int(articles["stock"].sum()) if not articles.empty else 0)

    render_report_download_box()

    if st.session_state.get("missing_name_alert"):
        render_big_missing_name_alert()

    st.divider()

    left, right = st.columns([1, 1])

    with left:
        personne = st.text_input("Nom de la personne qui complète", placeholder="Ex : Lucas")
        if personne.strip():
            st.session_state.missing_name_alert = False

    with right:
        recherche = st.text_input("Recherche article", placeholder="Ex : Pepsi, Gin, Champagne...")

    if recherche.strip():
        recherche_clean = recherche.strip().lower()
        articles_affiches = articles[
            articles["article"].str.lower().str.startswith(recherche_clean, na=False)
        ].copy()
    else:
        articles_affiches = articles.copy()

    if articles_affiches.empty:
        st.info("Aucun article trouvé.")
    else:
        h1, h2, h3, h4 = st.columns([3.2, 1, 2, 2], gap="small")
        with h1:
            st.markdown('<div class="table-head">Article</div>', unsafe_allow_html=True)
        with h2:
            st.markdown('<div class="table-head">Stock</div>', unsafe_allow_html=True)
        with h3:
            st.markdown('<div class="table-head">Sortie</div>', unsafe_allow_html=True)
        with h4:
            st.markdown('<div class="table-head">Entrée</div>', unsafe_allow_html=True)

        # Les quantités sont saisies dans un formulaire Streamlit.
        # Avantage : cliquer sur + / - dans les champs ne relance plus toute l'app.
        # Supabase n'est appelé qu'une seule fois, au moment de la validation.
        with st.form("mouvements_form", clear_on_submit=False):
            movements = []

            for i, row in articles_affiches.reset_index(drop=True).iterrows():
                cls = "cell-a" if i % 2 == 0 else "cell-b"
                article_id = int(row["id"])

                c1, c2, c3, c4 = st.columns([3.2, 1, 2, 2], gap="small")

                with c1:
                    st.markdown(
                        f'<div class="{cls}"><div class="article">{row["article"]}</div></div>',
                        unsafe_allow_html=True
                    )

                with c2:
                    st.markdown(
                        f'<div class="{cls}"><div class="stock">{int(row["stock"])}</div></div>',
                        unsafe_allow_html=True
                    )

                with c3:
                    sortie = st.number_input(
                        "Sortie",
                        min_value=0,
                        step=1,
                        value=int(st.session_state.get(movement_key("sortie", article_id), 0) or 0),
                        key=movement_key("sortie", article_id),
                        label_visibility="collapsed",
                    )

                with c4:
                    entree = st.number_input(
                        "Entrée",
                        min_value=0,
                        step=1,
                        value=int(st.session_state.get(movement_key("entree", article_id), 0) or 0),
                        key=movement_key("entree", article_id),
                        label_visibility="collapsed",
                    )

                movements.append({
                    "id": article_id,
                    "article": row["article"],
                    "stock": int(row["stock"]),
                    "sortie": int(sortie or 0),
                    "entree": int(entree or 0),
                })

                st.markdown('<div class="row-gap"></div>', unsafe_allow_html=True)

            st.divider()

            submitted = st.form_submit_button(
                "✅ Valider les mouvements",
                type="primary",
                use_container_width=True,
            )

        if submitted:
            if not personne.strip():
                st.session_state.missing_name_alert = True
                render_big_missing_name_alert()
                st.error("Validation impossible : le nom de la personne est obligatoire.")
            else:
                st.session_state.missing_name_alert = False
                count = 0
                mouvements_valides = []

                for move in movements:
                    if move["sortie"] or move["entree"]:
                        new_stock = move["stock"] - move["sortie"] + move["entree"]
                        mouvements_valides.append({
                            "id": move["id"],
                            "article": move["article"],
                            "sortie": move["sortie"],
                            "entree": move["entree"],
                            "ancien_stock": move["stock"],
                            "nouveau_stock": new_stock,
                        })
                        count += 1

                if count:
                    apply_movements(personne.strip(), mouvements_valides)

                if count == 0:
                    st.info("Aucune sortie ou entrée renseignée.")
                else:
                    try:
                        prepare_download_report(personne.strip(), mouvements_valides)
                        reset_movements()
                        st.rerun()
                    except Exception as e:
                        st.warning(
                            "Stock mis à jour, mais le rapport Excel n'a pas pu être généré. "
                            f"Erreur : {e}"
                        )

    low_stock = articles[articles["stock"] <= seuil].copy()

    if not low_stock.empty:
        low_stock = low_stock.sort_values(
            by=["stock", "article"],
            ascending=[True, True],
            key=lambda col: col.str.lower() if col.name == "article" else col
        )

        st.divider()
        st.subheader("Stock faible")

        st.dataframe(
            low_stock[["article", "stock"]],
            use_container_width=True,
            hide_index=True,
            height=280
        )


elif st.session_state.page == "resume":
    st.subheader("Stats mensuelles par article")

    mouvements = get_mouvements()

    if mouvements.empty:
        st.info("Aucun mouvement enregistré pour le moment.")
    else:
        mouvements["date_dt"] = pd.to_datetime(mouvements["date_jour"])
        mouvements["mois"] = mouvements["date_dt"].dt.strftime("%Y-%m")
        mouvements["mois_affichage"] = mouvements["date_dt"].dt.strftime("%m/%Y")

        mois_disponibles = (
            mouvements[["mois", "mois_affichage"]]
            .drop_duplicates()
            .sort_values("mois", ascending=False)
        )

        col_month, col_search = st.columns([1, 2])

        with col_month:
            mois_choisi = st.selectbox(
                "Mois",
                mois_disponibles["mois_affichage"].tolist()
            )

        with col_search:
            recherche_stats = st.text_input(
                "Rechercher un article",
                placeholder="Ex : Smirnoff, Champagne, Pepsi..."
            )

        mois_key = mois_disponibles.loc[
            mois_disponibles["mois_affichage"] == mois_choisi,
            "mois"
        ].iloc[0]

        filtered = mouvements[mouvements["mois"] == mois_key].copy()

        if recherche_stats.strip():
            filtered = filtered[
                filtered["article"].str.contains(recherche_stats.strip(), case=False, na=False)
            ]

        if filtered.empty:
            st.info("Aucun mouvement sur ce mois.")
        else:
            by_article = (
                filtered
                .groupby("article", as_index=False)
                .agg({
                    "sortie": "sum",
                    "entree": "sum"
                })
            )

            by_article["solde"] = by_article["entree"] - by_article["sortie"]

            by_article = by_article.sort_values(
                by=["sortie", "article"],
                ascending=[False, True],
                key=lambda col: col.str.lower() if col.name == "article" else col
            )

            total_sorties = int(by_article["sortie"].sum())
            total_entrees = int(by_article["entree"].sum())
            article_top = by_article.iloc[0]["article"] if not by_article.empty else "-"

            m1, m2, m3 = st.columns(3)
            m1.metric("Sorties du mois", total_sorties)
            m2.metric("Entrées du mois", total_entrees)
            m3.metric("Article le plus sorti", article_top)

            st.divider()

            st.markdown("### Classement des sorties du mois")

            classement = by_article.rename(columns={
                "article": "Article",
                "sortie": "Sorties",
                "entree": "Entrées",
                "solde": "Solde"
            })

            st.dataframe(
                classement[["Article", "Sorties", "Entrées", "Solde"]],
                use_container_width=True,
                hide_index=True,
                height=520
            )

            st.divider()

            st.markdown("### Détail du mois")

            detail = filtered.sort_values(
                ["date_jour", "heure"],
                ascending=[False, False]
            ).rename(columns={
                "date_jour": "Date",
                "heure": "Heure",
                "personne": "Personne",
                "article": "Article",
                "sortie": "Sortie",
                "entree": "Entrée",
                "ancien_stock": "Ancien stock",
                "nouveau_stock": "Nouveau stock",
            })

            detail["Date"] = pd.to_datetime(detail["Date"]).dt.strftime("%d/%m/%Y")

            st.dataframe(
                detail[[
                    "Date", "Heure", "Personne", "Article",
                    "Sortie", "Entrée", "Ancien stock", "Nouveau stock"
                ]],
                use_container_width=True,
                hide_index=True,
                height=360
            )


elif st.session_state.page == "admin":
    st.subheader("🔒 Admin")

    password = st.text_input("Mot de passe admin", type="password")

    if password != ADMIN_PASSWORD:
        st.info("Entre le mot de passe admin.")
    else:
        st.success("Accès admin autorisé")

        st.markdown("### Réglages")

        st.session_state.seuil_stock_faible = st.number_input(
            "Seuil stock faible",
            min_value=0,
            max_value=100,
            value=int(st.session_state.seuil_stock_faible),
            step=1
        )

        st.divider()

        st.markdown("### Modifier le stock / supprimer un article")

        articles_admin = get_articles()
        recherche_admin = st.text_input("Recherche admin", placeholder="Ex : Smirnoff")

        if recherche_admin.strip():
            articles_admin = articles_admin[
                articles_admin["article"].str.contains(recherche_admin.strip(), case=False, na=False)
            ]

        admin_view = articles_admin.copy()
        admin_view["supprimer"] = False

        edited_admin = st.data_editor(
            admin_view,
            column_config={
                "id": None,
                "article": st.column_config.TextColumn("Article", disabled=True, width="large"),
                "stock": st.column_config.NumberColumn("Stock", step=1, width="medium"),
                "supprimer": st.column_config.CheckboxColumn("Supprimer -"),
            },
            disabled=["article"],
            hide_index=True,
            use_container_width=True,
            height=520,
            key="table_admin"
        )

        to_delete = edited_admin[edited_admin["supprimer"] == True]

        if not to_delete.empty:
            st.markdown(
                '<div class="delete-note">Suppression demandée : '
                + ", ".join(to_delete["article"].tolist())
                + "</div>",
                unsafe_allow_html=True
            )

        save_col, delete_col = st.columns(2)

        with save_col:
            if st.button("💾 Enregistrer les stocks", use_container_width=True):
                for _, row in edited_admin.iterrows():
                    update_stock(row["id"], row["stock"])

                st.success("Stocks enregistrés.")
                st.rerun()

        with delete_col:
            if st.button("🗑️ Supprimer les articles cochés", use_container_width=True):
                if to_delete.empty:
                    st.info("Aucun article coché.")
                else:
                    for _, row in to_delete.iterrows():
                        delete_article(row["id"])

                    st.success("Article(s) supprimé(s).")
                    st.rerun()

        st.divider()

        st.markdown("### Ajouter un article")

        add_col1, add_col2 = st.columns([3, 1])

        with add_col1:
            new_article = st.text_input("Nom de l'article", placeholder="Ex : Vodka XYZ 70cl")

        with add_col2:
            new_stock = st.number_input("Stock initial", min_value=-999, value=0, step=1)

        if st.button("➕ Ajouter l'article", use_container_width=True):
            if not new_article.strip():
                st.error("Écris le nom de l'article.")
            else:
                try:
                    add_article(new_article, new_stock)
                    st.success("Article ajouté.")
                    st.rerun()
                except (sqlite3.IntegrityError, SQLAlchemyIntegrityError):
                    st.error("Cet article existe déjà.")